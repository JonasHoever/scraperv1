from flask import Flask, render_template, request, jsonify, flash, make_response, session
import os
from dotenv import load_dotenv
import logging
import pandas as pd
from io import BytesIO
from datetime import datetime
import json
from werkzeug.utils import secure_filename
from utils.geocoding import get_coordinates, search_insurance_brokers
from utils.scraper import scrape_broker_website
from utils.api_client import forward_to_external_api

# Umgebungsvariablen laden
load_dotenv()

app = Flask(__name__)
app.secret_key = os.getenv('SECRET_KEY', 'fallback_secret_key_for_development')

# Upload-Konfiguration
UPLOAD_FOLDER = 'uploads'
ALLOWED_EXTENSIONS = {'xls', 'xlsx'}
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size

# Upload-Ordner erstellen falls nicht vorhanden
if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)

# Logging konfigurieren
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


def allowed_file(filename):
    """Überprüft ob die Datei-Extension erlaubt ist"""
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def parse_excel_file(filepath):
    """Liest Excel-Datei und extrahiert Makler-Daten"""
    try:
        # Excel-Datei lesen
        df = pd.read_excel(filepath)
        
        # Erwartet: Spalten wie 'Name', 'Address', 'Phone', 'Email', 'Website', etc.
        brokers = []
        
        for index, row in df.iterrows():
            broker = {}
            
            # Standard-Spalten versuchen zu mappen
            possible_name_cols = ['Name', 'Firmenname', 'Company', 'Unternehmen', 'Makler']
            possible_address_cols = ['Address', 'Adresse', 'Straße', 'Street']
            possible_phone_cols = ['Phone', 'Telefon', 'Tel', 'Mobile']
            possible_email_cols = ['Email', 'E-Mail', 'Mail']
            possible_website_cols = ['Website', 'URL', 'Homepage', 'Web']
            
            # Name finden
            for col in possible_name_cols:
                if col in df.columns and pd.notna(row[col]):
                    broker['name'] = str(row[col]).strip()
                    break
            
            # Adresse finden
            for col in possible_address_cols:
                if col in df.columns and pd.notna(row[col]):
                    broker['address'] = str(row[col]).strip()
                    break
            
            # Telefon finden
            for col in possible_phone_cols:
                if col in df.columns and pd.notna(row[col]):
                    broker['phone'] = str(row[col]).strip()
                    break
            
            # Email finden
            for col in possible_email_cols:
                if col in df.columns and pd.notna(row[col]):
                    broker['email'] = str(row[col]).strip()
                    break
            
            # Website finden
            for col in possible_website_cols:
                if col in df.columns and pd.notna(row[col]):
                    broker['website'] = str(row[col]).strip()
                    break
            
            # Nur hinzufügen wenn mindestens Name vorhanden
            if 'name' in broker and broker['name']:
                brokers.append(broker)
        
        return brokers
        
    except Exception as e:
        logger.error(f"Fehler beim Lesen der Excel-Datei: {str(e)}")
        return None


def find_duplicates(existing_brokers, new_brokers):
    """Findet Duplikate zwischen bestehenden und neuen Maklern"""
    duplicates = []
    unique_new = []
    
    for new_broker in new_brokers:
        is_duplicate = False
        
        # Prüfung auf Duplikate basierend auf Name oder Adresse
        for existing in existing_brokers:
            if ('name' in new_broker and 'name' in existing and 
                new_broker['name'].lower() == existing['name'].lower()) or \
               ('address' in new_broker and 'address' in existing and 
                new_broker['address'].lower() == existing['address'].lower()):
                is_duplicate = True
                duplicates.append(new_broker)
                break
        
        if not is_duplicate:
            unique_new.append(new_broker)
    
    return unique_new, duplicates


@app.route('/')
def index():
    """Hauptseite mit Suchformular"""
    return render_template('index.html')


@app.route('/search', methods=['POST'])
def search_brokers():
    """Suche nach Versicherungsmaklern basierend auf Standort und Radius"""
    try:
        # Formulardaten auslesen
        location = request.form.get('location', '').strip()
        radius_km = int(request.form.get('radius', 10))
        
        if not location:
            flash('Bitte geben Sie eine Postleitzahl oder einen Ort ein.', 'error')
            return render_template('index.html')
        
        if radius_km < 1 or radius_km > 100:
            flash('Der Radius muss zwischen 1 und 100 km liegen.', 'error')
            return render_template('index.html')
        
        logger.info(f"Suche nach Versicherungsmaklern in {location} im Umkreis von {radius_km}km")
        
        # Koordinaten des Standorts ermitteln
        coordinates = get_coordinates(location)
        if not coordinates:
            flash('Standort konnte nicht gefunden werden. Bitte überprüfen Sie die Eingabe.', 'error')
            return render_template('index.html')
        
        # Versicherungsmakler in der Nähe suchen
        brokers = search_insurance_brokers(coordinates, radius_km * 1000)  # Umwandlung km zu m
        
        if not brokers:
            flash('Keine Versicherungsmakler in der angegebenen Region gefunden.', 'info')
            return render_template('index.html')
        
        # Details für jeden Makler durch Web-Scraping ergänzen
        enhanced_brokers = []
        for broker in brokers[:10]:  # Limitierung auf 10 Ergebnisse für Performance
            try:
                # Web-Scraping für zusätzliche Details
                scraped_data = scrape_broker_website(broker.get('website', ''))
                
                # Daten zusammenführen
                enhanced_broker = {
                    'name': broker.get('name', 'Unbekannt'),
                    'address': broker.get('formatted_address', 'Unbekannt'),
                    'phone': broker.get('formatted_phone_number', scraped_data.get('phone', 'Nicht verfügbar')),
                    'website': broker.get('website', 'Nicht verfügbar'),
                    'email': scraped_data.get('email', 'Nicht verfügbar'),
                    'contact_person': scraped_data.get('contact_person', 'Nicht verfügbar'),
                    'rating': broker.get('rating', 0),
                    'user_ratings_total': broker.get('user_ratings_total', 0),
                    'place_id': broker.get('place_id', ''),
                    'search_location': location,
                    'search_radius': radius_km,
                    'found_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                }
                
                enhanced_brokers.append(enhanced_broker)
                
            except Exception as e:
                logger.warning(f"Fehler beim Scraping für {broker.get('name', 'Unbekannt')}: {str(e)}")
                # Fallback ohne Scraping-Daten
                enhanced_broker = {
                    'name': broker.get('name', 'Unbekannt'),
                    'address': broker.get('formatted_address', 'Unbekannt'),
                    'phone': broker.get('formatted_phone_number', 'Nicht verfügbar'),
                    'website': broker.get('website', 'Nicht verfügbar'),
                    'email': 'Nicht verfügbar',
                    'contact_person': 'Nicht verfügbar',
                    'rating': broker.get('rating', 0),
                    'user_ratings_total': broker.get('user_ratings_total', 0),
                    'place_id': broker.get('place_id', ''),
                    'search_location': location,
                    'search_radius': radius_km,
                    'found_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                }
                enhanced_brokers.append(enhanced_broker)
        
        # Suchergebnisse in Session speichern für Excel-Export
        session['last_search_results'] = enhanced_brokers
        session['last_search_params'] = {
            'location': location,
            'radius': radius_km,
            'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        }
        
        logger.info(f"Gefunden: {len(enhanced_brokers)} Versicherungsmakler")
        
        return render_template('results.html', 
                             brokers=enhanced_brokers, 
                             location=location, 
                             radius=radius_km)
        
    except ValueError:
        flash('Ungültiger Radius. Bitte geben Sie eine Zahl ein.', 'error')
        return render_template('index.html')
    except Exception as e:
        logger.error(f"Fehler bei der Maklersuche: {str(e)}")
        flash('Ein Fehler ist aufgetreten. Bitte versuchen Sie es später erneut.', 'error')
        return render_template('index.html')


@app.route('/api/brokers', methods=['GET'])
def api_get_brokers():
    """API Endpoint um alle gefundenen Makler als JSON zu erhalten"""
    # Hier könnte man eine Datenbank abfragen oder Session-Daten verwenden
    # Für diese Demo geben wir ein Beispiel zurück
    return jsonify({
        'status': 'success',
        'message': 'Verwenden Sie das Suchformular, um Makler zu finden.',
        'brokers': []
    })


@app.route('/api/test', methods=['GET', 'POST'])
def api_test_connection():
    """Test-Endpoint um die externe API-Verbindung zu testen"""
    from utils.api_client import test_external_api_connection, prepare_broker_payload
    
    if request.method == 'GET':
        # Zeige Test-Interface
        return render_template('api_test.html')
    
    try:
        # Test-Daten erstellen
        test_broker = {
            'name': 'Test Versicherungsmakler GmbH',
            'contact_person': 'Max Mustermann',
            'address': 'Teststraße 123, 12345 Teststadt',
            'phone': '+49 123 456789',
            'email': 'test@example.com',
            'website': 'https://test-makler.de',
            'rating': 4.5,
            'user_ratings_total': 42,
            'place_id': 'test_place_id_12345'
        }
        
        # Test-Payload erstellen
        payload = prepare_broker_payload(test_broker)
        
        # An externe API senden
        response = forward_to_external_api(test_broker)
        
        if response and response.get('success'):
            return jsonify({
                'status': 'success',
                'message': 'API-Test erfolgreich',
                'sent_payload': payload,
                'api_response': response,
                'format_used': os.getenv('API_SEND_FORMAT', 'enhanced')
            })
        else:
            return jsonify({
                'status': 'error',
                'message': 'API-Test fehlgeschlagen',
                'sent_payload': payload,
                'error_details': response,
                'format_used': os.getenv('API_SEND_FORMAT', 'enhanced')
            }), 500
            
    except Exception as e:
        logger.error(f"Fehler beim API-Test: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': f'API-Test Fehler: {str(e)}'
        }), 500


@app.route('/api/settings', methods=['GET', 'POST'])
def api_settings():
    """API-Einstellungen verwalten"""
    if request.method == 'GET':
        # Aktuelle Einstellungen anzeigen
        current_settings = {
            'external_api_url': os.getenv('EXTERNAL_API_URL', ''),
            'api_send_format': os.getenv('API_SEND_FORMAT', 'enhanced'),
            'external_api_key': os.getenv('EXTERNAL_API_KEY', ''),
            'external_api_token': os.getenv('EXTERNAL_API_TOKEN', ''),
            'timeout': os.getenv('API_TIMEOUT', '30'),
            'retry_attempts': os.getenv('API_RETRY_ATTEMPTS', '3')
        }
        return render_template('api_settings.html', settings=current_settings)
    
    elif request.method == 'POST':
        try:
            # Einstellungen aus Formular lesen
            new_settings = {
                'EXTERNAL_API_URL': request.form.get('external_api_url', '').strip(),
                'API_SEND_FORMAT': request.form.get('api_send_format', 'enhanced'),
                'EXTERNAL_API_KEY': request.form.get('external_api_key', '').strip(),
                'EXTERNAL_API_TOKEN': request.form.get('external_api_token', '').strip(),
                'API_TIMEOUT': request.form.get('timeout', '30'),
                'API_RETRY_ATTEMPTS': request.form.get('retry_attempts', '3')
            }
            
            # Validierung
            if new_settings['EXTERNAL_API_URL']:
                from urllib.parse import urlparse
                parsed = urlparse(new_settings['EXTERNAL_API_URL'])
                if not parsed.scheme or not parsed.netloc:
                    flash('Ungültige API-URL. Bitte eine vollständige URL eingeben (z.B. https://api.example.com/webhook)', 'error')
                    return render_template('api_settings.html', settings=request.form)
            
            # Timeout validierung
            try:
                timeout_val = int(new_settings['API_TIMEOUT'])
                if timeout_val < 5 or timeout_val > 300:
                    raise ValueError()
            except ValueError:
                flash('Timeout muss zwischen 5 und 300 Sekunden liegen.', 'error')
                return render_template('api_settings.html', settings=request.form)
            
            # .env Datei aktualisieren oder erstellen
            env_file_path = '.env'
            env_content = {}
            
            # Bestehende .env laden falls vorhanden
            if os.path.exists(env_file_path):
                with open(env_file_path, 'r', encoding='utf-8') as f:
                    for line in f:
                        line = line.strip()
                        if line and not line.startswith('#') and '=' in line:
                            key, value = line.split('=', 1)
                            env_content[key.strip()] = value.strip()
            
            # Neue Werte setzen
            for key, value in new_settings.items():
                if value:  # Nur nicht-leere Werte setzen
                    env_content[key] = value
                elif key in env_content:  # Leere Werte entfernen
                    del env_content[key]
            
            # .env Datei schreiben
            with open(env_file_path, 'w', encoding='utf-8') as f:
                f.write('# Versicherungsmakler Finder - API Konfiguration\n')
                f.write('# Automatisch generiert am ' + datetime.now().strftime('%Y-%m-%d %H:%M:%S') + '\n\n')
                
                for key, value in env_content.items():
                    f.write(f'{key}={value}\n')
            
            # Umgebungsvariablen neu laden
            from dotenv import load_dotenv
            load_dotenv(override=True)
            
            flash('API-Einstellungen erfolgreich gespeichert! Die Änderungen sind sofort aktiv.', 'success')
            logger.info(f"API-Einstellungen aktualisiert: URL={new_settings.get('EXTERNAL_API_URL', 'nicht gesetzt')}")
            
            return render_template('api_settings.html', settings=new_settings)
            
        except Exception as e:
            logger.error(f"Fehler beim Speichern der API-Einstellungen: {str(e)}")
            flash('Fehler beim Speichern der Einstellungen. Bitte versuchen Sie es erneut.', 'error')
            return render_template('api_settings.html', settings=request.form)


@app.route('/api/test-connection', methods=['POST'])
def test_api_connection():
    """Test der aktuellen API-Verbindung"""
    try:
        # Test-Daten erstellen
        test_data = {
            'test': True,
            'timestamp': datetime.now().isoformat(),
            'source': 'Versicherungsmakler Finder - Connection Test',
            'data': {
                'name': 'Test Versicherungsmakler GmbH',
                'contact_person': 'Max Mustermann',
                'address': 'Teststraße 123, 12345 Teststadt',
                'phone': '+49 123 456789',
                'email': 'test@beispiel-makler.de',
                'website': 'https://beispiel-makler.de'
            }
        }
        
        # Test-Request senden
        api_url = os.getenv('EXTERNAL_API_URL')
        if not api_url:
            return jsonify({
                'status': 'error',
                'message': 'Keine API-URL konfiguriert'
            }), 400
        
        response = forward_to_external_api(test_data)
        
        return jsonify({
            'status': 'success' if response and response.get('success') else 'error',
            'api_url': api_url,
            'response': response,
            'test_data': test_data
        })
        
    except Exception as e:
        logger.error(f"API-Verbindungstest fehlgeschlagen: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': f'Verbindungstest fehlgeschlagen: {str(e)}'
        }), 500
    """API Endpoint um Makler-Daten an externe API weiterzuleiten"""
    try:
        broker_data = request.get_json()
        
        if not broker_data:
            return jsonify({
                'status': 'error',
                'message': 'Keine Daten empfangen'
            }), 400
        
        # An externe API weiterleiten
        response = forward_to_external_api(broker_data)
        
        if response and response.get('success'):
            return jsonify({
                'status': 'success',
                'message': 'Daten erfolgreich weitergeleitet',
                'external_response': response,
                'format_used': os.getenv('API_SEND_FORMAT', 'enhanced')
            })
        else:
            return jsonify({
                'status': 'error',
                'message': 'Fehler beim Weiterleiten der Daten',
                'error_details': response
            }), 500
            
    except Exception as e:
        logger.error(f"Fehler beim Weiterleiten der API-Daten: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': 'Interner Serverfehler'
        }), 500


@app.route('/export/excel', methods=['GET'])
def export_excel():
    """Excel-Export der letzten Suchergebnisse"""
    try:
        # Suchergebnisse aus Session abrufen
        brokers = session.get('last_search_results', [])
        search_params = session.get('last_search_params', {})
        
        if not brokers:
            flash('Keine Suchergebnisse zum Export verfügbar. Führen Sie zuerst eine Suche durch.', 'warning')
            return render_template('index.html')
        
        # DataFrame erstellen
        df = pd.DataFrame(brokers)
        
        # Spaltennamen auf Deutsch übersetzen und sortieren
        column_mapping = {
            'name': 'Name',
            'contact_person': 'Ansprechpartner',
            'address': 'Adresse',
            'phone': 'Telefon',
            'email': 'E-Mail',
            'website': 'Website',
            'rating': 'Bewertung',
            'user_ratings_total': 'Anzahl Bewertungen',
            'search_location': 'Suchort',
            'search_radius': 'Suchradius (km)',
            'found_at': 'Gefunden am',
            'place_id': 'Google Place ID'
        }
        
        df = df.rename(columns=column_mapping)
        
        # Spalten in gewünschter Reihenfolge
        desired_columns = [
            'Name', 'Ansprechpartner', 'Adresse', 'Telefon', 
            'E-Mail', 'Website', 'Bewertung', 'Anzahl Bewertungen',
            'Suchort', 'Suchradius (km)', 'Gefunden am', 'Google Place ID'
        ]
        
        df = df[desired_columns]
        
        # Excel-Datei im Speicher erstellen
        output = BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Hauptdaten
            df.to_excel(writer, sheet_name='Versicherungsmakler', index=False)
            
            # Suchparameter auf separatem Arbeitsblatt
            search_info = pd.DataFrame([{
                'Parameter': 'Suchort',
                'Wert': search_params.get('location', 'Unbekannt')
            }, {
                'Parameter': 'Suchradius (km)',
                'Wert': search_params.get('radius', 'Unbekannt')
            }, {
                'Parameter': 'Suchdatum',
                'Wert': search_params.get('timestamp', 'Unbekannt')
            }, {
                'Parameter': 'Anzahl Ergebnisse',
                'Wert': len(brokers)
            }])
            
            search_info.to_excel(writer, sheet_name='Suchparameter', index=False)
            
            # Arbeitsblätter formatieren
            workbook = writer.book
            
            # Makler-Arbeitsblatt formatieren
            worksheet = writer.sheets['Versicherungsmakler']
            
            # Spaltenbreite anpassen
            column_widths = {
                'A': 25,  # Name
                'B': 20,  # Ansprechpartner
                'C': 40,  # Adresse
                'D': 15,  # Telefon
                'E': 25,  # E-Mail
                'F': 30,  # Website
                'G': 10,  # Bewertung
                'H': 15,  # Anzahl Bewertungen
                'I': 20,  # Suchort
                'J': 15,  # Suchradius
                'K': 20,  # Gefunden am
                'L': 25   # Place ID
            }
            
            for col, width in column_widths.items():
                worksheet.column_dimensions[col].width = width
            
            # Header-Formatierung
            from openpyxl.styles import Font, PatternFill, Alignment
            
            header_font = Font(bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
            
            # Daten formatieren
            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
                for cell in row:
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
        
        output.seek(0)
        
        # Response erstellen
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        location_clean = search_params.get('location', 'Suche').replace(' ', '_').replace(',', '')
        filename = f'Versicherungsmakler_{location_clean}_{timestamp}.xlsx'
        
        response = make_response(output.read())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename={filename}'
        
        logger.info(f"Excel-Export erstellt: {filename}")
        return response
        
    except Exception as e:
        logger.error(f"Fehler beim Excel-Export: {str(e)}")
        flash('Fehler beim Erstellen der Excel-Datei. Bitte versuchen Sie es erneut.', 'error')
        return render_template('index.html')


@app.route('/export/json', methods=['GET'])
def export_json():
    """JSON-Export der letzten Suchergebnisse"""
    try:
        # Suchergebnisse aus Session abrufen
        brokers = session.get('last_search_results', [])
        search_params = session.get('last_search_params', {})
        
        if not brokers:
            return jsonify({
                'status': 'error',
                'message': 'Keine Suchergebnisse verfügbar'
            }), 400
        
        export_data = {
            'metadata': {
                'export_timestamp': datetime.now().isoformat(),
                'total_results': len(brokers),
                'search_params': search_params,
                'app_version': '2.0',
                'format_version': '1.0'
            },
            'brokers': brokers
        }
        
        # JSON-Response mit Download-Header
        response = make_response(jsonify(export_data))
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        location_clean = search_params.get('location', 'Suche').replace(' ', '_').replace(',', '')
        filename = f'Versicherungsmakler_{location_clean}_{timestamp}.json'
        
        response.headers['Content-Type'] = 'application/json'
        response.headers['Content-Disposition'] = f'attachment; filename={filename}'
        
        logger.info(f"JSON-Export erstellt: {filename}")
        return response
        
    except Exception as e:
        logger.error(f"Fehler beim JSON-Export: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': f'Fehler beim JSON-Export: {str(e)}'
        }), 500


@app.route('/upload', methods=['GET', 'POST'])
def upload_excel():
    """Excel-Upload für Makler-Listen und erweiterte Suche"""
    if request.method == 'GET':
        return render_template('upload.html')
    
    try:
        # Datei-Upload prüfen
        if 'file' not in request.files:
            flash('Keine Datei ausgewählt!', 'error')
            return render_template('upload.html')
        
        file = request.files['file']
        
        if file.filename == '':
            flash('Keine Datei ausgewählt!', 'error')
            return render_template('upload.html')
        
        if not allowed_file(file.filename):
            flash('Nur Excel-Dateien (.xls, .xlsx) sind erlaubt!', 'error')
            return render_template('upload.html')
        
        # Suchparameter aus Formular
        location = request.form.get('location', '').strip()
        radius_km = int(request.form.get('radius', 10))
        
        if not location:
            flash('Bitte geben Sie eine Postleitzahl oder einen Ort für die erweiterte Suche ein.', 'error')
            return render_template('upload.html')
        
        # Datei sicher speichern
        filename = secure_filename(file.filename)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        unique_filename = f"{timestamp}_{filename}"
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], unique_filename)
        file.save(filepath)
        
        logger.info(f"Excel-Datei hochgeladen: {unique_filename}")
        
        # Excel-Datei parsen
        existing_brokers = parse_excel_file(filepath)
        
        if existing_brokers is None:
            flash('Fehler beim Lesen der Excel-Datei. Bitte überprüfen Sie das Format.', 'error')
            os.remove(filepath)  # Temporäre Datei löschen
            return render_template('upload.html')
        
        if not existing_brokers:
            flash('Keine gültigen Makler-Daten in der Excel-Datei gefunden.', 'warning')
            os.remove(filepath)
            return render_template('upload.html')
        
        logger.info(f"{len(existing_brokers)} Makler aus Excel-Datei gelesen")
        
        # Koordinaten des Suchstandorts ermitteln
        coordinates = get_coordinates(location)
        if not coordinates:
            flash('Suchstandort konnte nicht gefunden werden. Bitte überprüfen Sie die Eingabe.', 'error')
            os.remove(filepath)
            return render_template('upload.html')
        
        # Neue Makler in der Zone suchen
        logger.info(f"Suche nach zusätzlichen Maklern in {location} im Umkreis von {radius_km}km")
        new_brokers_raw = search_insurance_brokers(coordinates, radius_km * 1000)
        
        if not new_brokers_raw:
            flash('Keine neuen Makler in der angegebenen Zone gefunden.', 'info')
            results = {
                'existing_count': len(existing_brokers),
                'new_count': 0,
                'duplicate_count': 0,
                'existing_brokers': existing_brokers,
                'new_brokers': [],
                'duplicates': [],
                'search_location': location,
                'search_radius': radius_km
            }
            os.remove(filepath)
            return render_template('upload_results.html', results=results)
        
        logger.info(f"{len(new_brokers_raw)} neue Makler in der Zone gefunden")
        
        # Duplikate filtern
        unique_new_brokers, duplicates = find_duplicates(existing_brokers, new_brokers_raw)
        
        logger.info(f"{len(unique_new_brokers)} neue einzigartige Makler gefunden, {len(duplicates)} Duplikate")
        
        # Detaillierte Informationen für neue Makler scrapen
        enhanced_new_brokers = []
        for i, broker in enumerate(unique_new_brokers[:10]):  # Limit für Performance
            try:
                logger.info(f"Scraping Makler {i+1}/{min(len(unique_new_brokers), 10)}: {broker.get('name', 'Unbekannt')}")
                enhanced_data = scrape_broker_website(broker.get('website', ''))
                if enhanced_data:
                    broker.update(enhanced_data)
                enhanced_new_brokers.append(broker)
            except Exception as e:
                logger.warning(f"Scraping für {broker.get('name', 'Unbekannt')} fehlgeschlagen: {str(e)}")
                enhanced_new_brokers.append(broker)
        
        # Ergebnisse in Session speichern für Export
        session['upload_results'] = {
            'existing_brokers': existing_brokers,
            'new_brokers': enhanced_new_brokers,
            'duplicates': duplicates,
            'search_params': {
                'location': location,
                'radius': radius_km,
                'timestamp': datetime.now().isoformat()
            }
        }
        
        # Temporäre Datei löschen
        os.remove(filepath)
        
        results = {
            'existing_count': len(existing_brokers),
            'new_count': len(enhanced_new_brokers),
            'duplicate_count': len(duplicates),
            'existing_brokers': existing_brokers,
            'new_brokers': enhanced_new_brokers,
            'duplicates': duplicates,
            'search_location': location,
            'search_radius': radius_km
        }
        
        flash(f'Upload erfolgreich! {len(existing_brokers)} bestehende, {len(enhanced_new_brokers)} neue Makler gefunden.', 'success')
        return render_template('upload_results.html', results=results)
        
    except Exception as e:
        logger.error(f"Fehler beim Excel-Upload: {str(e)}")
        flash(f'Fehler beim Upload: {str(e)}', 'error')
        return render_template('upload.html')


@app.errorhandler(404)
def not_found(error):
    return render_template('404.html'), 404


@app.errorhandler(500)
def internal_error(error):
    return render_template('500.html'), 500


if __name__ == '__main__':
    # Überprüfung der wichtigsten Umgebungsvariablen
    if not os.getenv('GOOGLE_MAPS_API_KEY'):
        logger.warning("GOOGLE_MAPS_API_KEY nicht gesetzt! Bitte .env Datei konfigurieren.")
    
    app.run(debug=os.getenv('FLASK_DEBUG', 'False').lower() == 'true', 
            host='0.0.0.0', 
            port=int(os.getenv('PORT', 5000)))
